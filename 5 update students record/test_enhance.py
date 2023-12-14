import openpyxl
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

def test():
   # try:
   driver=webdriver.Edge()
   driver.get("http://127.0.0.1:5500/")
   students = [
      ["John", "27", "Male", "B+" ],
      ["amy", "31", "Female", "A" ],
      ["Bob", "23", "Male", "B" ],
      ["Alice", "28", "Female", "A" ],
      ["Charli", "22", "Male", "C" ],
      ["David", "25", "Male", "B+" ],
      ["Eva", "30", "Female", "A" ],
      ["Frank", "26", "Male", "A+" ],
      ["Grace", "29", "Female", "B" ],
      ["Hannah", "24", "Female", "C+" ]
   ]
   student_data = []
   for student in students:
      name_input = driver.find_element(By.ID, 'name')
      name_input.send_keys(student[0])

      age_input = driver.find_element(By.ID, "age")
      age_input.send_keys(student[1])

      gen_input = driver.find_element(By.ID, "gender")
      gen_input.send_keys(student[2])

      grade_input = driver.find_element(By.ID, "grade")
      grade_input.send_keys(student[3])

      update_btn = driver.find_element(By.XPATH, '//*[@id="updateForm"]/input[5]')
      update_btn.click()
      sleep(1)

      msg = driver.find_element(By.ID, "msg").text
      print(f"msg={msg} type={type(msg)}")
      if msg == "student record updated successfully!":
         print(f"student:{student[0]} updated successfully")
      else:
         print("A wrong input was given aborting updates")
         return

   table = driver.find_element(By.ID, "myTable")
   rows = table.find_elements(By.TAG_NAME, "tr")
   for row in rows[1:]:
      cells = row.find_elements(By.TAG_NAME, "td")
      name = cells[0].text
      age = cells[1].text
      gen = cells[2].text
      grade = cells[3].text
      student_data.append([name, age, gen, grade])

   print("Updating...")
   wb = openpyxl.Workbook()
   ws =wb.active
   ws.cell(row=1, column=1).value = "Name:"
   ws.cell(row=1, column=2).value = "Age:"
   ws.cell(row=1, column=3).value = "Gender:"
   ws.cell(row=1, column=4).value = "Grade:"

   for i, data in enumerate(student_data, start=2):
      print(f"i={i}\t name={data[0]}\t age={data[1]}\t gen={data[2]}\t grade={data[3]}")
      ws.cell(row=i, column=1).value = data[0]
      ws.cell(row=i, column=2).value = data[1]
      ws.cell(row=i, column=3).value = data[2]
      ws.cell(row=i, column=4).value = data[3]

   wb.save("output_student.xlsx")
   print("successfully added to sheet")
   driver.quit()

   # except Exception as e:
   #    print("Error...",e)
test()