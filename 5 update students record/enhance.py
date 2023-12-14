import openpyxl
from time import sleep
from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

def test():
   try:
      driver=webdriver.Edge()
      driver.get("http://127.0.0.1:5500/templates/")
      students = [["amy","20","F","A+"]]
      student_data = []
      for student in students:
         name_input = driver.find_element(By.ID, "name")
         name_input.send_keys(student[0])

         age_input = driver.find_element(By.ID, "age")
         age_input.send_keys(student[1])

         gen_input = driver.find_element(By.ID, "gender")
         gen_input.send_keys(student[2])

         grade_input = driver.find_element(By.ID, "grade")
         grade_input.send_keys(student[3])
         sleep(3)
         update_btn = driver.find_element(By.XPATH, '//*[@id="updateForm"]/input[5]')
         update_btn.click()

         msg = driver.find_element(By.ID, "msg").text
         print(f"msg={msg} type={type(msg)}")
         if msg == "student record updated successfully!":
            print(f"student:{student[0]} updated")
         else:
            print("A wrong input was given aborting updates")
            return
         student_data.append(student)
      # table = driver.find_element(By.ID, "myTable")
      # rows = table.find_elements(By.TAG_NAME, "tr")
      # for row in rows:
      #    cells = row.find_elements(By.TAG_NAME, "td")
      #    name = cells[0].text
      #    age = cells[1].text
      #    gen = cells[2].text
      #    grade = cells[3].text
      #    student_data.append([name], [age], [gen], [grade])
      driver.quit()

      print("Updating...")
      wb = openpyxl.Workbook()
      ws =wb.active
      ws.cell(row=1, column=1).value = "name:"
      ws.cell(row=1, column=2).value = "age:"
      ws.cell(row=1, column=3).value = "gender"
      ws.cell(row=1, column=4).value = "grade"

      for i, data in enumerate(student_data):
         ws.cell(row=i+2, column=1).value = data[0]
         ws.cell(row=i+2, column=2).value = data[1]
         ws.cell(row=i+2, column=3).value = data[2]
         ws.cell(row=i+2, column=4).value = data[3]
      wb.save("output_pr5_students.xlsx")
      print("successfully added to sheet")
   except Exception as e:
      print("Error...",e)
test()